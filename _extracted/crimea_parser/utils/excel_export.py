"""CSV → XLSX с вкладками по городам, форматированием, кликабельными контактами.

Структура книги:
  «Сводка»       — статистика по источникам/городам/типам клиентов
  «Все»          — полный список (фильтр + поиск)
  <Город>        — отдельный лист на каждый город из CSV
  «Без контактов» — записи без phone и без email (для ручного добивания)

Форматирование:
  - заморожена шапка
  - AutoFilter, ширина по содержимому, перенос текста для address/comment
  - заливка строки по client_type
  - кликабельные phone (tel:), email (mailto:), website
"""
from __future__ import annotations

import csv
import os
from collections import Counter, defaultdict
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


HEADERS = [
    "Город", "Название", "Тип клиента", "Категория",
    "Адрес", "Телефон", "Email", "Сайт", "Соцсеть",
    "Комментарий", "Источник", "Собрано",
]

# Соответствие CSV-полей → колонки Excel
CSV_FIELD_ORDER = [
    "city", "name", "client_type", "category",
    "address", "phone", "email", "website", "social",
    "comment", "source", "parsed_at",
]

# Заливка строки по client_type
FILL_BY_TYPE = {
    "отель":         "DCEEFB",  # голубой
    "гостиница":     "DCEEFB",
    "апарт-отель":   "E2D9F3",
    "апартаменты":   "E2D9F3",
    "пансионат":     "DCFCE7",  # зелёный
    "санаторий":     "FEF3C7",  # жёлтый
    "база отдыха":   "FFF7ED",
    "дом отдыха":    "FFF7ED",
    "гостевой дом":  "FEF3C7",
    "хостел":        "F3F4F6",  # серый
    "эллинг":        "FFE4E6",
    "глэмпинг":      "DCFCE7",
    "кемпинг":       "DCFCE7",
    "мотель":        "F3F4F6",
    "прочее":        "FFFFFF",
}

# Заливка для шапки и стилевые заголовки
HEADER_FILL = PatternFill("solid", fgColor="1F2937")  # тёмно-серый
HEADER_FONT = Font(bold=True, color="FFFFFF")

# Цвет источника (для бейджа в колонке Источник)
SOURCE_FONT_COLOR = {
    "OSM": "1E40AF",         # синий
    "Я.Карты": "B91C1C",     # красный
    "2ГИС": "047857",        # зелёный
    "Авито": "C2410C",       # оранжевый
    "Суточно.ру": "9333EA",  # фиолетовый
    "Ostrovok": "0F766E",    # бирюзовый
    "Wikidata": "374151",    # серый
}

NO_CONTACT_FILL = PatternFill("solid", fgColor="FECACA")  # светло-красный

WRAP_COLS = {"Адрес", "Комментарий"}


def _read_csv(path: str) -> list[dict]:
    rows = []
    for delim in (";", ","):
        try:
            with open(path, encoding="utf-8-sig") as f:
                sample = f.read(2048)
                f.seek(0)
                if delim in sample:
                    reader = csv.DictReader(f, delimiter=delim)
                    rows = list(reader)
                    if rows and "name" in rows[0]:
                        return rows
        except Exception:
            continue
    return rows


def _safe_sheet_name(name: str) -> str:
    """Excel запрещает : \\ / ? * [ ] в именах листов и >31 символа."""
    bad = set(":\\/?*[]")
    out = "".join(" " if c in bad else c for c in name)
    return out.strip()[:31] or "Лист"


def _write_sheet(ws, rows: list[dict], with_filter: bool = True) -> None:
    """Запись данных в лист с форматированием."""
    # шапка
    for col_idx, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")

    if not rows:
        ws.freeze_panes = "A2"
        return

    # данные
    for r_idx, row in enumerate(rows, start=2):
        client_type = (row.get("client_type") or "прочее").lower()
        fill_color = FILL_BY_TYPE.get(client_type, "FFFFFF")
        no_phone = not (row.get("phone") or "").strip()
        no_email = not (row.get("email") or "").strip()
        row_fill = NO_CONTACT_FILL if (no_phone and no_email) else (
            PatternFill("solid", fgColor=fill_color) if fill_color != "FFFFFF" else None)

        for c_idx, csv_field in enumerate(CSV_FIELD_ORDER, start=1):
            val = row.get(csv_field, "") or ""
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            header = HEADERS[c_idx - 1]

            if header in WRAP_COLS:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
            else:
                cell.alignment = Alignment(vertical="top")

            if row_fill is not None:
                cell.fill = row_fill

            # Гиперссылки
            if header == "Сайт" and val.startswith(("http://", "https://")):
                cell.hyperlink = val
                cell.font = Font(color="0563C1", underline="single")
            elif header == "Email" and "@" in val:
                cell.hyperlink = f"mailto:{val}"
                cell.font = Font(color="0563C1", underline="single")
            elif header == "Телефон" and val:
                tel = "".join(ch for ch in val if ch.isdigit() or ch == "+")
                if tel:
                    cell.hyperlink = f"tel:{tel}"
                    cell.font = Font(color="0563C1", underline="single")
            elif header == "Соцсеть" and val.startswith(("http://", "https://")):
                cell.hyperlink = val
                cell.font = Font(color="0563C1", underline="single")
            elif header == "Источник":
                color = SOURCE_FONT_COLOR.get(val, "374151")
                cell.font = Font(bold=True, color=color)

    # ширина колонок
    widths = {
        "Город": 14, "Название": 36, "Тип клиента": 14, "Категория": 16,
        "Адрес": 40, "Телефон": 20, "Email": 28, "Сайт": 30, "Соцсеть": 24,
        "Комментарий": 30, "Источник": 14, "Собрано": 16,
    }
    for c_idx, header in enumerate(HEADERS, start=1):
        ws.column_dimensions[get_column_letter(c_idx)].width = widths.get(header, 16)

    ws.freeze_panes = "A2"
    if with_filter:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{len(rows) + 1}"

    # высота строки шапки
    ws.row_dimensions[1].height = 22


def _write_summary(ws, rows: list[dict], src_csv: str) -> None:
    """Сводка: общая статистика."""
    total = len(rows)
    with_phone = sum(1 for r in rows if (r.get("phone") or "").strip())
    with_email = sum(1 for r in rows if (r.get("email") or "").strip())
    with_addr = sum(1 for r in rows if (r.get("address") or "").strip())
    with_site = sum(1 for r in rows if (r.get("website") or "").strip())
    with_social = sum(1 for r in rows if (r.get("social") or "").strip())

    src_cnt = Counter(r.get("source", "—") for r in rows)
    city_cnt = Counter(r.get("city", "—") for r in rows)
    type_cnt = Counter(r.get("client_type", "—") for r in rows)

    def pct(n: int) -> str:
        return f"{100 * n / total:.1f}%" if total else "—"

    lines: list[tuple[str, str]] = [
        ("Файл", os.path.basename(src_csv)),
        ("Сформирован", datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("", ""),
        ("Всего записей", str(total)),
        ("С телефоном", f"{with_phone} ({pct(with_phone)})"),
        ("С email", f"{with_email} ({pct(with_email)})"),
        ("С адресом", f"{with_addr} ({pct(with_addr)})"),
        ("С сайтом", f"{with_site} ({pct(with_site)})"),
        ("С соцсетью", f"{with_social} ({pct(with_social)})"),
        ("", ""),
        ("Источники", ""),
    ]
    for k, v in src_cnt.most_common():
        lines.append((f"  {k}", str(v)))
    lines.append(("", ""))
    lines.append(("Типы клиентов", ""))
    for k, v in type_cnt.most_common():
        lines.append((f"  {k}", str(v)))
    lines.append(("", ""))
    lines.append(("Города (топ 20)", ""))
    for k, v in city_cnt.most_common(20):
        lines.append((f"  {k}", str(v)))

    bold = Font(bold=True)
    for r_idx, (k, v) in enumerate(lines, start=1):
        ws.cell(row=r_idx, column=1, value=k).font = bold if not k.startswith("  ") else Font()
        ws.cell(row=r_idx, column=2, value=v)
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 30


def build_xlsx(csv_path: str, xlsx_path: str | None = None) -> str | None:
    """Построить XLSX из CSV. Возвращает путь к файлу."""
    if not os.path.exists(csv_path):
        print(f"[xlsx] CSV не найден: {csv_path}")
        return None

    rows = _read_csv(csv_path)
    if not rows:
        print(f"[xlsx] CSV пуст: {csv_path}")
        return None

    if not xlsx_path:
        base = os.path.splitext(csv_path)[0]
        xlsx_path = base + ".xlsx"

    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "Сводка"
    _write_summary(ws_summary, rows, csv_path)

    ws_all = wb.create_sheet("Все")
    _write_sheet(ws_all, rows)

    # Группа по городам — отдельный лист на каждый
    by_city: dict[str, list[dict]] = defaultdict(list)
    for r in rows:
        city = (r.get("city") or "Не указан").strip() or "Не указан"
        by_city[city].append(r)

    for city in sorted(by_city.keys(), key=lambda c: -len(by_city[c])):
        sheet_name = _safe_sheet_name(city)
        if sheet_name in wb.sheetnames:
            sheet_name = _safe_sheet_name(f"{city}_2")
        ws = wb.create_sheet(sheet_name)
        _write_sheet(ws, by_city[city])

    # Без контактов
    no_contacts = [r for r in rows
                   if not (r.get("phone") or "").strip()
                   and not (r.get("email") or "").strip()]
    if no_contacts:
        ws_nc = wb.create_sheet("Без контактов")
        _write_sheet(ws_nc, no_contacts)

    os.makedirs(os.path.dirname(xlsx_path) or ".", exist_ok=True)
    wb.save(xlsx_path)
    print(f"[xlsx] сохранён: {xlsx_path} (листов: {len(wb.sheetnames)})")
    return xlsx_path
