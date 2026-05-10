import os
from datetime import date, datetime
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, colors
)
from openpyxl.utils import get_column_letter
from config import EXCEL_HEADERS

REPORTS_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)), "reports")


def _style_header_row(ws):
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=10)
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for col_num, header in enumerate(EXCEL_HEADERS, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border


def _style_data_row(ws, row_num: int, num_cols: int):
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    fill_even = PatternFill(start_color="EBF3FB", end_color="EBF3FB", fill_type="solid")
    fill_odd = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    fill = fill_even if row_num % 2 == 0 else fill_odd
    for col_num in range(1, num_cols + 1):
        cell = ws.cell(row=row_num, column=col_num)
        cell.fill = fill
        cell.border = border
        cell.alignment = Alignment(vertical="center", wrap_text=True)


def _set_column_widths(ws):
    widths = [5, 18, 30, 18, 16, 18, 25, 18, 25, 20, 25, 22, 30, 15, 35, 20, 20]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _leads_to_rows(leads: list) -> list:
    rows = []
    for idx, lead in enumerate(leads, 1):
        rows.append([
            idx,
            lead.get("created_at", ""),
            lead.get("name", ""),
            lead.get("object_type", ""),
            lead.get("city", ""),
            lead.get("region", ""),
            lead.get("address", ""),
            lead.get("phone", ""),
            lead.get("email", ""),
            lead.get("telegram", ""),
            lead.get("website", ""),
            lead.get("size", ""),
            lead.get("interests", ""),
            lead.get("status", ""),
            lead.get("comment", ""),
            lead.get("added_by", ""),
            lead.get("status_updated_at", ""),
        ])
    return rows


def _build_workbook(leads: list, sheet_title: str) -> openpyxl.Workbook:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_title
    ws.row_dimensions[1].height = 30

    _style_header_row(ws)

    rows = _leads_to_rows(leads)
    for row_num, row_data in enumerate(rows, 2):
        for col_num, value in enumerate(row_data, 1):
            ws.cell(row=row_num, column=col_num, value=value)
        _style_data_row(ws, row_num, len(EXCEL_HEADERS))

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"
    _set_column_widths(ws)
    return wb


def generate_daily_report(leads: list) -> str:
    os.makedirs(REPORTS_DIR, exist_ok=True)
    today_str = date.today().strftime("%Y-%m-%d")
    filename = f"daily_{today_str}.xlsx"
    filepath = os.path.join(REPORTS_DIR, filename)
    wb = _build_workbook(leads, f"Лиды {today_str}")
    wb.save(filepath)
    return filepath


def generate_weekly_report(leads: list) -> str:
    os.makedirs(REPORTS_DIR, exist_ok=True)
    today_str = date.today().strftime("%Y-%m-%d")
    filename = f"weekly_{today_str}.xlsx"
    filepath = os.path.join(REPORTS_DIR, filename)
    wb = _build_workbook(leads, f"Неделя до {today_str}")
    wb.save(filepath)
    return filepath


def generate_master_report(leads: list) -> str:
    os.makedirs(REPORTS_DIR, exist_ok=True)
    today_str = date.today().strftime("%Y-%m-%d")
    filename = f"master_leads_{today_str}.xlsx"
    filepath = os.path.join(REPORTS_DIR, filename)
    wb = _build_workbook(leads, "Все контакты")
    wb.save(filepath)
    return filepath
